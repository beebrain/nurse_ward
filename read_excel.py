import pandas as pd
import openpyxl

file = "6. ยอดรายวันและ Productivity มี.ค. 68.xlsx"

wb = openpyxl.load_workbook(file, read_only=False, data_only=False)
names = wb.sheetnames

def get_productivity_formula(ws, sheet_name):
    """Extract BF/BG formula from first data row of any ward sheet."""
    print(f"\n=== {sheet_name}: Productivity formulas (row 4) ===")
    for col in range(54, 68):
        cell = ws.cell(row=4, column=col)
        if cell.value:
            print(f"  {cell.coordinate}: {str(cell.value)[:200]}")
    # Also get header row labels
    for col in range(54, 68):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            print(f"  header {cell.coordinate}: {cell.value}")

# Check formula in several different ward types
for sname in ['ICUM', 'CCU', 'Stroke Unit', 'RCW', 'EENT']:
    if sname in names:
        get_productivity_formula(wb[sname], sname)
    else:
        # Try garbled name
        for n in names:
            if sname.lower() in n.lower() or n.lower() in sname.lower():
                get_productivity_formula(wb[n], n)
                break

# Check non-ICU ward - try sheet index 11 (อย1?)
print(f"\n=== Sheet[11] ({names[11]}): Productivity formulas ===")
ws11 = wb[names[11]]
for col in range(54, 68):
    cell = ws11.cell(row=4, column=col)
    if cell.value:
        print(f"  {cell.coordinate}: {str(cell.value)[:200]}")
for col in range(54, 68):
    cell = ws11.cell(row=1, column=col)
    if cell.value:
        print(f"  header {cell.coordinate}: {cell.value}")

# Look at Night/summary per ward statistics area (summary rows at bottom)
print(f"\n=== ICUM: Bottom rows (rows 95-105) ===")
ws_icum = wb['ICUM']
for row in ws_icum.iter_rows(min_row=95, max_row=105, values_only=False):
    for cell in row:
        if cell.value is not None and str(cell.value).strip():
            print(f"  {cell.coordinate}: {str(cell.value)[:150]}")

wb.close()
